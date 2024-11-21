import vt
import os
import hashlib
import openpyxl
import json
from jinja2 import Environment, FileSystemLoader

# Constants for configuration
API_KEY_PATH = "api_key.txt"
TEMPLATE_DIRECTORY = 'templates'
REPORT_TEMPLATE_NAME = 'hash_report_template.md'
HASH_LIST_FILENAME = "hashes.xlsx"
GENERATED_REPORTS_DIR = "reports"
API_RESPONSES_DIR = "responses"

class ReportGenerator:
    def __init__(self):
        # Initialize Jinja2 environment and load the template
        self.env = Environment(loader=FileSystemLoader(TEMPLATE_DIRECTORY))
        self.template = self.env.get_template(REPORT_TEMPLATE_NAME)

    def generate(self, response):
        # Render the template with the given response data
        return self.template.render(
            meaningful_name=response.get('meaningful_name'),
            label=response.get('popular_threat_classification', {}).get('suggested_threat_label'),
            reputation=response.get('reputation'),
            sandbox_verdicts=response.get('sandbox_verdicts'),
            total_votes=response.get('total_votes')
        )

def generate_hash_list_from_folder(folder_path, xlsx_filename):
    # Create an Excel file with filenames and their MD5 hashes
    wb = openpyxl.Workbook()
    ws = wb.active

    for i, filename in enumerate(os.listdir(folder_path), start=1):
        with open(os.path.join(folder_path, filename), "rb") as file:
            ws.cell(row=i, column=1).value = filename
            ws.cell(row=i, column=2).value = hashlib.md5(file.read()).hexdigest()

    wb.save(xlsx_filename)

def fetch_report_from_virustotal(client, file_hash):
    # Fetch the report of a file from VirusTotal using its hash
    return client.get_object(f"/files/{file_hash}")

def read_json_file(filename):
    # Load data from a JSON file
    with open(filename, "r") as f:
        return json.load(f)

def extract_hashes_from_excel(xlsx_filename):
    # Extract file hashes from an Excel file
    wb = openpyxl.load_workbook(xlsx_filename)
    ws = wb.active
    return [ws.cell(row=i, column=2).value for i in range(1, ws.max_row + 1)]

def save_data(filename, data, is_json_format=False):
    # Save data to a file, either as JSON or plain text
    try:
        with open(filename, "w") as f:
            if is_json_format:
                # Convert custom objects to a regular dict if necessary
                if isinstance(data, vt.Object):
                    data = data.to_dict()
                json.dump(data, f, indent=4)
            else:
                f.write(data)
    except TypeError as e:
        print(f"Serialization error: {e}. File: {filename}")

def retrieve_api_key(filename):
    # Load the API key from a file
    with open(filename, 'r') as f:
        return f.read().strip()

def main():
    api_key = retrieve_api_key(API_KEY_PATH)

    with vt.Client(api_key) as client:
        hashes = extract_hashes_from_excel(HASH_LIST_FILENAME)
        report_gen = ReportGenerator()

        for h in hashes:
            try:
                print(f"Processing hash: {h}")
                response = fetch_report_from_virustotal(client, h)
                response_dict = response.to_dict()  # Convert response to dict
                save_data(f"{API_RESPONSES_DIR}/{h}.json", response_dict, is_json_format=True)
                report_content = report_gen.generate(response_dict)
                save_data(f"{GENERATED_REPORTS_DIR}/{h}.md", report_content)
                print(f"Report generated for hash: {h}")
            except vt.APIError as e:
                print(f"Error with hash {h}: {e}")
            except Exception as e:
                print(f"Unexpected error with hash {h}: {e}")

if __name__ == "__main__":
    main()
